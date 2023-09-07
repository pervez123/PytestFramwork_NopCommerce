import inspect
import random
import string
import time
from datetime import datetime
import openpyxl
import logging
import pytest
from selenium.common import NoSuchElementException, exceptions
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait


@pytest.mark.usefixtures("setup")
class BaseClass:

    customer_menu_link = (By.XPATH, "//li/a[@href='#']//p[contains(text(),'Customers')]")
    customer_sub_menu_customer = (By.XPATH, "//a[@href='/Admin/Customer/List']//p[contains(text(),'Customers')]")
    customer_sub_menu_customer_roles = (By.XPATH, "//a[@href='/Admin/CustomerRole/List']//p"
                                                  "[contains(text(),'Customer roles')]")
    customer_sub_menu_vendor = (By.XPATH, "//a[@href=\"/Admin/Vendor/List\"]/p[contains(text(),'Vendors')]")
    customer_sub_menu_online_customer = (By.XPATH, "//a[@href='/Admin/OnlineCustomer/List']//"
                                                   "p[contains(text(),'Online customer')]")
    catalog_menu_link = (By.XPATH, "//a[@href='#']//p[contains(text(),'Catalog')]")
    catalog_manufacturer = (By.XPATH, "//a[@href='/Admin/Manufacturer/List']//p[text() = ' Manufacturers']")
    catalog_submenu_product = (By.XPATH, "//a[@href='/Admin/Product/List']//p[text() = ' Products']")
    links = (By.TAG_NAME, "a")
    btn_save = (By.NAME, "save")
    alert_success = (By.XPATH, "//div[contains(@class, 'alert-success')]")
    alert_error = (By.XPATH, "//div[contains(@class, 'alert-danger')]")
    locator_field_validation_error = (By.CSS_SELECTOR, ".field-validation-error")
    locator_summary_validation_error = (By.XPATH, "//div[@class = 'validation-summary-errors']//li")

    def find_element(self, *locator):
        return self.driver.find_element(*locator)

    def find_elements(self, *locator):
        return self.driver.find_elements(*locator)

    def click_element(self, *locator):
        self.find_element(*locator).click()

    def set_element(self, locator, value):
        # self.find_element(*locator).send_keys(value)
        element = self.wait_until_presence_of_element_located(locator)
        element.send_keys(value)

    def clear_element(self, *locator):
        self.find_element(*locator).clear()

    def get_text(self, *locator):
        return self.find_element(*locator).text

    def action_method(self):
        action = ActionChains(self.driver)
        return action

    def scroll_to_element(self, locator):
        element = self.find_element(*locator)
        self.driver.execute_script("arguments[0].scrollIntoView();", element)

    def scroll_to_end(self):
        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

    def calender(self, calender_month_year_locator, calender_next_icon_locator, date_locator, travel_month_year,
                 travel_date):
        # Current month year showing on the cal.
        parsed_date = datetime.strptime(travel_month_year, '%B %Y')
        current_month_year = self.get_text(*calender_month_year_locator)
        while current_month_year != travel_month_year:
            # clicking next icon given on cal.
            self.wait_until_element_to_be_clickable(calender_next_icon_locator).click()
            # Fetching month year after clicking on next icon.
            current_month_year = self.get_text(*calender_month_year_locator)
        time.sleep(1)
        date = self.find_elements(*date_locator)
        for i in date:
            if i.text == travel_date:
                time.sleep(1)
                i.click()
                break
        selected_date = self.driver.find_element(By.ID, "PreOrderAvailabilityStartDateTimeUtc").get_attribute('value')
        travel_formatted_date = parsed_date.strftime(f"%m/{travel_date}/%Y")
        assert travel_formatted_date in selected_date

    def click_save_btn(self):
        self.click_element(*self.btn_save)

    def click_on_customer_menu_link(self):
        self.click_element(*self.customer_menu_link)

    def click_on_customer_sub_menu_link(self):
        self.click_element(*self.customer_sub_menu_customer)

    def click_on_vendor_sub_menu_link(self):
        self.click_element(*self.customer_sub_menu_vendor)

    def click_on_customer_roles_sub_menu_link(self):
        self.click_element(*self.customer_sub_menu_customer_roles)

    def click_on_online_customer_sub_menu_link(self):
        self.click_element(*self.customer_sub_menu_online_customer)

    def select_value_from_static_list(self, field_locator, list_locator, value):
        self.click_element(*field_locator)
        time.sleep(1)
        # static_list = self.find_elements(*list_locator)
        static_list = self.wait_until_presence_of_all_elements_located(list_locator)
        for i in static_list:
            if i.text == value:
                i.click()
                break

    def wait_until_element_to_be_clickable(self, *locator):
        wait = WebDriverWait(self.driver, 10)
        return wait.until(EC.element_to_be_clickable(*locator))

    def wait_until_presence_of_element_located(self, *locator):
        wait = WebDriverWait(self.driver, 10)
        return wait.until(EC.presence_of_element_located(*locator))

    def wait_until_presence_of_all_elements_located(self, *locator):
        wait = WebDriverWait(self.driver, 10)
        return wait.until(EC.presence_of_all_elements_located(*locator))

    def wait_for_text(self, locator, text):
        wait = WebDriverWait(self.driver, 10)
        return wait.until(EC.text_to_be_present_in_element(locator, text))

    def select_by_visible_text(self, locator, text):
        select = Select(self.find_element(*locator))
        select.select_by_visible_text(text)

    def logs(self):
        test_filename = inspect.stack()[1][3]
        logger = logging.getLogger(test_filename)
        file_handler = logging.FileHandler(".\\Logs\\logfile.log")  # Defining location of the log file
        formatter = logging.Formatter('%(asctime)s : %(levelname)s : %(name)s : %(message)s')
        file_handler.setFormatter(formatter)
        if logger.hasHandlers():
            logger.handlers.clear()
        logger.addHandler(file_handler)
        logger.addHandler(file_handler)
        logger.setLevel(logging.INFO)
        return logger

    @staticmethod
    def excel_test_data(filename, test_case_name):
        book = openpyxl.load_workbook(filename)
        sheet = book.active
        rows = sheet.max_row
        columns = sheet.max_column
        lst = []
        for i in range(1, rows+1):
            dic = {}
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, columns+1):
                    dic[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                lst.append(dic)
        return lst

    def random_email_generator(self, size=8, chars=string.ascii_lowercase + string.digits):
        return ''.join(random.choice(chars) for i in range(size))

    def success_alert_pop_up(self):
        return self.find_element(*self.alert_success)

    def error_alert_pop_up(self):
        return self.find_element(*self.alert_error)

    def field_validation(self):
        return self.find_element(*self.locator_field_validation_error)

    def summary_validation(self):
        return self.find_element(*self.locator_summary_validation_error)

    def field_validation_error_message(self):
        return self.get_text(*self.locator_field_validation_error)

    def success_alert(self):
        return self.get_text(*self.alert_success)

    def error_alert(self):
        self.get_text(*self.alert_error)

    def validation(self, dataset, back_to_list=()):
        lst = []
        try:
            success_message = self.success_alert_pop_up()

            if success_message:
                if dataset['exp_result'] == 'pass':
                    lst.append("pass")
                    self.logs().info(f"-----Test case passed with message **{self.success_alert()}**-----------")

                elif dataset['exp_result'] == 'fail':
                    lst.append('fail')
                    self.logs().error(f"-----Test case failed with message **{self.success_alert()}**-----------")

        except NoSuchElementException:
            try:
                error_message = self.error_alert_pop_up()
                if error_message:
                    if dataset['exp_result'] == 'pass':
                        lst.append("fail")
                        self.logs().error(f"-----Test case failed with message **{self.error_alert()}**----------")

                    elif dataset['exp_result'] == 'fail':
                        lst.append('pass')
                        self.logs().info(f"-----Test case passed with message **{self.error_alert()}**-----------")

            except NoSuchElementException:
                try:
                    validation_error_message = self.field_validation()
                    if validation_error_message:
                        time.sleep(2)
                        if dataset['exp_result'] == 'pass':
                            lst.append("fail")
                            self.logs().error(
                                f"---Test case failed with message **{self.field_validation_error_message()}**---")
                            try:
                                self.click_element(*back_to_list)
                            except exceptions:
                                pass

                        elif dataset['exp_result'] == 'fail':
                            lst.append('pass')
                            self.logs().info(
                                f"--Test case passed with message **{self.field_validation_error_message()}**--")
                            try:
                                self.click_element(*back_to_list)
                            except exceptions:
                                pass
                except NoSuchElementException:
                    summary_validation_error = self.summary_validation()
                    if summary_validation_error:
                        time.sleep(2)
                        if dataset['exp_result'] == 'pass':
                            lst.append("fail")
                            self.logs().error(
                                "---Test case failed with message **"
                                f"{self.get_text(*self.locator_summary_validation_error)}**---")
                            try:
                                self.click_element(*back_to_list)
                            except exceptions:
                                pass

                        elif dataset['exp_result'] == 'fail':
                            lst.append('pass')
                            self.logs().info(
                                "--Test case passed with message **"
                                f"{self.get_text(*self.locator_summary_validation_error)}**--")
                            try:
                                self.click_element(*back_to_list)
                            except exceptions:
                                pass

        if "fail" in lst:
            self.logs().error("********* Some of the combination failed ***********")
            assert False

        else:
            assert True

